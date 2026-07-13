import os
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import openpyxl

# Nombre del archivo de Excel
EXCEL_FILE = "Reporte de Asistencia.xlsx"
SHEET_NAME = "Listado de aprendices"

def registrar_datos():
    # 1. Obtener los datos de los campos de texto
    documento = entry_doc.get().strip()
    correo = entry_correo.get().strip()
    celular = entry_celular.get().strip()
    
    # 2. Validaciones básicas
    if not documento or not correo or not celular:
        messagebox.showerror("Error", "Todos los campos son obligatorios.")
        return
    
    # 3. Obtener fecha y hora local automáticamente
    fecha_hora_local = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 4. Verificar existencia del archivo Excel
    if not os.path.exists(EXCEL_FILE):
        messagebox.showerror("Error", f"No se encontró el archivo '{EXCEL_FILE}' en el directorio actual.")
        return

    try:
        # Abrir el libro y la hoja usando openpyxl para mantener formatos
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if SHEET_NAME not in wb.sheetnames:
            messagebox.showerror("Error", f"No se encontró la hoja '{SHEET_NAME}' en el archivo.")
            wb.close()
            return
            
        sheet = wb[SHEET_NAME]
        
        # Mapeo de columnas solicitadas:
        # Columna L es la número 12 (Original: Número de documento)
        # Columna T = 20, U = 21, V = 22, W = 23
        col_L = 12
        col_T = 20
        col_U = 21
        col_V = 22
        col_W = 23
        
        coincidencia_encontrada = False
        
        # Recorrer las filas buscando la coincidencia (empezando desde la fila 2 para omitir encabezados)
        for row in range(2, sheet.max_row + 1):
            val_L = sheet.cell(row=row, column=col_L).value
            
            # Convertir a string para evitar líos si en Excel está guardado como número o texto
            if val_L is not None and str(val_L).strip() == str(documento):
                # Escribir en las columnas T, U, V, W respectivamente
                sheet.cell(row=row, column=col_T, value=fecha_hora_local)
                sheet.cell(row=row, column=col_U, value=str(documento))
                sheet.cell(row=row, column=col_V, value=correo)
                sheet.cell(row=row, column=col_W, value=celular)
                coincidencia_encontrada = True
                # No hacemos 'break' por si el mismo documento aparece registrado más de una vez
        
        if coincidencia_encontrada:
            wb.save(EXCEL_FILE)
            messagebox.showinfo("Éxito", f"Registro guardado exitosamente para el documento {documento}.")
            # Limpiar el formulario
            entry_doc.delete(0, tk.END)
            entry_correo.delete(0, tk.END)
            entry_celular.delete(0, tk.END)
        else:
            messagebox.showwarning("No encontrado", f"El número de documento '{documento}' no se encontró en la columna L.")
            
        wb.close()

    except Exception as e:
        messagebox.showerror("Error Técnico", f"Ocurrió un error al procesar el archivo:\n{e}")

# --- Configuración de la Ventana Principal (Tkinter) ---
root = tk.Tk()
root.title("Registro de Aprendices - SENA")
root.geometry("400x320")
root.resizable(False, False)

# Estilos y espaciado básicos
label_titulo = tk.Label(root, text="Formulario de Asistencia / Actualización", font=("Arial", 12, "bold"), pady=10)
label_titulo.pack()

frame_campos = tk.Frame(root, padx=20, pady=10)
frame_campos.pack(fill="both", expand=True)

# Campo: Número de Documento
lbl_doc = tk.Label(frame_campos, text="Número de Documento:", font=("Arial", 10))
lbl_doc.grid(row=0, column=0, sticky="w", pady=8)
entry_doc = tk.Entry(frame_campos, font=("Arial", 10), width=25)
entry_doc.grid(row=0, column=1, pady=8)

# Campo: Correo Electrónico
lbl_correo = tk.Label(frame_campos, text="Correo Electrónico:", font=("Arial", 10))
lbl_correo.grid(row=1, column=0, sticky="w", pady=8)
entry_correo = tk.Entry(frame_campos, font=("Arial", 10), width=25)
entry_correo.grid(row=1, column=1, pady=8)

# Campo: Número de Celular
lbl_celular = tk.Label(frame_campos, text="Número de Celular:", font=("Arial", 10))
lbl_celular.grid(row=2, column=0, sticky="w", pady=8)
entry_celular = tk.Entry(frame_campos, font=("Arial", 10), width=25)
entry_celular.grid(row=2, column=1, pady=8)

# Nota sobre la fecha
lbl_nota = tk.Label(root, text="* La fecha y hora local se registrarán automáticamente.", font=("Arial", 8, "italic"), fg="gray")
lbl_nota.pack(pady=5)

# Botón de Registro
btn_registrar = tk.Button(root, text="Guardar Registro", font=("Arial", 10, "bold"), bg="#238243", fg="white", command=registrar_datos, padding=6)
btn_registrar.pack(pady=15)

# Ejecutar la aplicación
root.mainloop()
