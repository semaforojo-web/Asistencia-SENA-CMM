import streamlit as st
from datetime import datetime
import openpyxl
import requests
import io
from github import Github

st.set_page_config(page_title="Registro de Aprendices - SENA", page_icon="📝")

st.title("Formulario de Asistencia / Actualización")

EXCEL_FILE = "Reporte de Asistencia.xlsx"
SHEET_NAME = "Listado de aprendices"
REPO_NAME = "semaforojo-web/Asistencia-SENA-CMM"  # Ajusta si tu usuario/repo cambia


# ==========================================
# FUNCIÓN AUXILIAR: DESCARGAR EXCEL DE GITHUB A MEMORIA
# ==========================================
def descargar_excel_desde_github():
    """Descarga el archivo Excel binario directamente usando la URL raw de GitHub
    para evitar corrupciones."""
    try:
        url_raw = f"https://raw.githubusercontent.com/{REPO_NAME}/main/{EXCEL_FILE}"

        headers = {}
        if "GITHUB_TOKEN" in st.secrets:
            headers = {"Authorization": f"token {st.secrets['GITHUB_TOKEN']}"}

        response = requests.get(url_raw, headers=headers)

        if response.status_code == 200:
            return io.BytesIO(response.content)
        else:
            st.error(f"⚠️ Error al acceder al archivo en GitHub (Código {response.status_code})")
    except Exception as e:
        st.error(f"⚠️ Fallo en la descarga directa: {e}")
    return None


# ==========================================
# FUNCIÓN: GUARDAR/SINCRONIZAR EL WORKBOOK MODIFICADO EN GITHUB
# ==========================================
def subir_contenido_a_github(content):
    """Sube el contenido binario (bytes) ya generado del Excel al repositorio
    de GitHub, actualizando el archivo si ya existe o creándolo si no.

    IMPORTANTE: el workbook de openpyxl debe guardarse UNA SOLA VEZ (con
    wb.save()) y reutilizar esos mismos bytes aquí y en el botón de
    descarga. Guardar el mismo Workbook más de una vez puede provocar
    'I/O operation on closed file' porque openpyxl cierra el archivo
    interno tras el primer save()."""
    if "GITHUB_TOKEN" not in st.secrets:
        st.warning("⚠️ No se detectó el 'GITHUB_TOKEN' en los Secrets de Streamlit.")
        return False

    try:
        g = Github(st.secrets["GITHUB_TOKEN"])
        repo = g.get_repo(REPO_NAME)

        try:
            contents = repo.get_contents(EXCEL_FILE, ref="main")
            sha = contents.sha
        except Exception:
            sha = None

        if sha:
            repo.update_file(
                path=EXCEL_FILE,
                message="🤖 Actualización de registro de asistencia",
                content=content,
                sha=sha,
                branch="main",
            )
        else:
            repo.create_file(
                path=EXCEL_FILE,
                message="🤖 Creación inicial del archivo de asistencia",
                content=content,
                branch="main",
            )
        return True
    except Exception as e:
        st.error(f"⚠️ Error crítico en la conexión con GitHub: {e}")
        return False


with st.form("formulario_asistencia", clear_on_submit=True):
    documento = st.text_input("Número de Documento:").strip()
    correo = st.text_input("Correo Electrónico:").strip()
    celular = st.text_input("Número de Celular:").strip()
    enviado = st.form_submit_button("Guardar Registro")

if enviado:
    if not documento or not correo or not celular:
        st.error("Todos los campos son obligatorios.")
    else:
        fecha_hora_local = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            archivo_memoria = descargar_excel_desde_github()
            if archivo_memoria is None:
                st.error("No se pudo descargar el archivo desde GitHub.")
            else:
                wb = openpyxl.load_workbook(archivo_memoria)
                sheet = wb[SHEET_NAME]
                coincidencia = False

                for row in range(2, sheet.max_row + 1):
                    val_L = sheet.cell(row=row, column=12).value
                    if val_L and str(val_L).split('.')[0].strip() == documento:
                        sheet.cell(row=row, column=20, value=fecha_hora_local)
                        sheet.cell(row=row, column=21, value=documento)
                        sheet.cell(row=row, column=22, value=correo)
                        sheet.cell(row=row, column=23, value=celular)
                        coincidencia = True

                if coincidencia:
                    # Guardamos el workbook UNA SOLA VEZ en memoria y reutilizamos
                    # estos mismos bytes tanto para GitHub como para la descarga.
                    output = io.BytesIO()
                    wb.save(output)
                    contenido_excel = output.getvalue()
                    wb.close()

                    subido = subir_contenido_a_github(contenido_excel)
                    if subido:
                        st.success("¡Registro guardado y sincronizado con GitHub!")

                        st.download_button(
                            label="Descargar Reporte Actualizado",
                            data=contenido_excel,
                            file_name="Reporte_Asistencia_Final.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                else:
                    st.warning("Documento no encontrado.")
                    wb.close()
        except Exception as e:
            st.error(f"Error técnico: {e}")
